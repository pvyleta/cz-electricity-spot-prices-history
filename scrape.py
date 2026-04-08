"""
Scraper for Czech electricity spot prices from OTE-CR.

Data sources (in order of preference):
  1. XLSX files from pubweb — used for recent/current data (2025-10-01+).
     These contain both 15min and 1h prices in a single download.
  2. JSON chart-data API — used as fallback for historical data (pre-2025).
     The same API the OTE website uses to render charts; works back to 2002.
     No XLSX parsing needed, zero extra dependencies.

Auto-detects resolution: tries 15min XLSX → 1h XLSX → JSON 1h.

Usage:
    # Scrape a single day
    python scrape.py --date 2026-04-08

    # Scrape a date range
    python scrape.py --from 2015-01-01 --to 2024-12-31

    # Scrape yesterday (for cron)
    python scrape.py --yesterday

Dependencies: openpyxl, requests
    pip install openpyxl requests
"""

import argparse
import csv
import io
import logging

from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
import requests

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)

BASE_URL = "https://www.ote-cr.cz/pubweb/attachments/01"
DATA_DIR = Path(__file__).parent / "data"

# XLSX layout constants (both formats share the same header row position)
HEADER_ROW = 22
# Row 22 = header, row 23 = empty, row 24 = first data row
DATA_START_ROW = 24


def xlsx_url_15min(d: date) -> str:
    return f"{BASE_URL}/{d.year}/month{d.month:02d}/day{d.day:02d}/DT_15MIN_{d.day:02d}_{d.month:02d}_{d.year}_CZ.xlsx"


def xlsx_url_1h(d: date) -> str:
    return f"{BASE_URL}/{d.year}/month{d.month:02d}/day{d.day:02d}/DT_{d.day:02d}_{d.month:02d}_{d.year}_CZ.xlsx"


def download_xlsx(url: str) -> bytes | None:
    """Download XLSX, return bytes or None if not found."""
    try:
        resp = requests.get(url, timeout=30)
        if resp.status_code == 200 and len(resp.content) > 500:
            return resp.content
        return None
    except requests.RequestException as e:
        log.warning("Request failed for %s: %s", url, e)
        return None


def parse_15min_xlsx(data: bytes, d: date) -> tuple[list[dict], list[dict]]:
    """Parse 15min XLSX. Returns (rows_15min, rows_1h).

    The 15min XLSX contains both 15min prices (col C) and 1h prices (col L).
    We extract both to avoid needing a separate 1h download.
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active

    rows_15min = []
    rows_1h = []
    seen_1h_intervals = set()  # deduplicate 1h prices (repeated per 15min slot)

    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        period_num = row[0]
        if period_num is None or not isinstance(period_num, (int, float)):
            break  # end of data (e.g. "Celkem" row)

        time_interval = row[1]  # e.g. "00:00-00:15"
        price_15min = row[2]    # 15min price EUR/MWh
        price_1h = row[11]      # 60min price EUR/MWh (col L, index 11)

        time_from, time_to = time_interval.split("-")

        rows_15min.append({
            "date": d.isoformat(),
            "time_from": time_from,
            "time_to": time_to,
            "price_eur_mwh": price_15min,
        })

        # 1h price is the same for every 15min slot within the hour — deduplicate
        if price_1h is not None and time_from not in seen_1h_intervals:
            # Derive the hour boundary from the 15min slot
            hour_start = time_from[:3] + "00"  # e.g. "21:00"
            hour_end_h = int(time_from[:2]) + 1
            hour_end = f"{hour_end_h:02d}:00" if hour_end_h < 24 else "24:00"

            if hour_start not in seen_1h_intervals:
                seen_1h_intervals.add(hour_start)
                rows_1h.append({
                    "date": d.isoformat(),
                    "time_from": hour_start,
                    "time_to": hour_end,
                    "price_eur_mwh": price_1h,
                })

    wb.close()
    return rows_15min, rows_1h


def parse_1h_xlsx(data: bytes, d: date) -> list[dict]:
    """Parse 1h-only XLSX (older format). Returns rows for 1h CSV."""
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        hour_num = row[0]
        if hour_num is None or not isinstance(hour_num, (int, float)):
            break

        hour = int(hour_num)
        price = row[1]  # EUR/MWh

        rows.append({
            "date": d.isoformat(),
            "time_from": f"{hour - 1:02d}:00",
            "time_to": f"{hour:02d}:00" if hour < 24 else "24:00",
            "price_eur_mwh": price,
        })

    wb.close()
    return rows


JSON_API = "https://www.ote-cr.cz/cs/kratkodobe-trhy/elektrina/denni-trh/@@chart-data"
CSV_HEADER = ["date", "time_from", "time_to", "price_eur_mwh"]


def scrape_day_json(d: date) -> list[dict]:
    """Fetch 1h prices via the JSON chart-data API (pre-2025 fallback).

    OTE's own charting API; available for all dates back to 2002.
    The response contains multiple dataLine objects — we select the one
    where useY2=False, which is always the price line (EUR/MWh).
    """
    params = {"report_date": d.isoformat(), "time_resolution": "PT60M"}
    try:
        resp = requests.get(JSON_API, params=params, timeout=30)
        if resp.status_code != 200:
            return []
        payload = resp.json()
    except Exception as e:
        log.warning("JSON API failed for %s: %s", d, e)
        return []

    # Find the price line — it's the only dataLine with useY2=False
    price_line = next(
        (dl for dl in payload["data"]["dataLine"] if not dl.get("useY2", True)),
        None,
    )
    if not price_line:
        log.warning("No price dataLine in JSON response for %s", d)
        return []

    rows = []
    for pt in price_line["point"]:
        hour = int(pt["x"])
        rows.append({
            "date": d.isoformat(),
            "time_from": f"{hour - 1:02d}:00",
            "time_to": f"{hour:02d}:00" if hour < 24 else "24:00",
            "price_eur_mwh": pt["y"],
        })
    return rows


def append_to_csv(filepath: Path, rows: list[dict]) -> None:
    """Append rows to a CSV file, creating it with header if needed."""
    filepath.parent.mkdir(parents=True, exist_ok=True)
    file_exists = filepath.exists() and filepath.stat().st_size > 0

    with open(filepath, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_HEADER)
        if not file_exists:
            writer.writeheader()
        writer.writerows(rows)


def date_already_scraped(filepath: Path, d: date) -> bool:
    """Check if data for a given date already exists in the CSV."""
    if not filepath.exists():
        return False
    with open(filepath, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row["date"] == d.isoformat():
                return True
    return False


# XLSX files are only published on pubweb from 2025 onwards.
# For earlier dates, skip straight to the JSON API to avoid wasted requests.
XLSX_AVAILABLE_FROM = date(2025, 1, 1)


def scrape_day(d: date) -> bool:
    """Scrape a single day. Returns True if any data was written."""
    csv_15min = DATA_DIR / "15min" / f"{d.year}.csv"
    csv_1h = DATA_DIR / "1h" / f"{d.year}.csv"

    # Skip if already scraped (check both — 15min may not exist for older dates)
    if date_already_scraped(csv_1h, d):
        log.debug("Already scraped %s, skipping", d)
        return False

    if d >= XLSX_AVAILABLE_FROM:
        # Try 15min first (available since 2025-10-01)
        url_15 = xlsx_url_15min(d)
        log.info("Trying 15min: %s", url_15)
        data = download_xlsx(url_15)

        if data:
            rows_15, rows_1h = parse_15min_xlsx(data, d)
            log.info("  %s: got %d 15min rows, %d 1h rows", d, len(rows_15), len(rows_1h))
            if rows_15:
                append_to_csv(csv_15min, rows_15)
            if rows_1h:
                append_to_csv(csv_1h, rows_1h)
            return True

        # Fall back to 1h-only XLSX
        url_1 = xlsx_url_1h(d)
        log.info("Trying 1h: %s", url_1)
        data = download_xlsx(url_1)

        if data:
            rows = parse_1h_xlsx(data, d)
            log.info("  %s: got %d 1h rows", d, len(rows))
            if rows:
                append_to_csv(csv_1h, rows)
            return True

    # JSON API: primary path for pre-2025, fallback for 2025+ if XLSX fails
    log.info("Trying JSON API for %s", d)
    rows = scrape_day_json(d)
    if rows:
        log.info("  %s: got %d 1h rows (JSON)", d, len(rows))
        append_to_csv(csv_1h, rows)
        return True

    log.warning("No data available for %s", d)
    return False


def scrape_range(start: date, end: date) -> None:
    """Scrape all days in [start, end] inclusive."""
    total_days = (end - start).days + 1
    scraped = 0
    skipped = 0
    failed = 0

    d = start
    while d <= end:
        try:
            if scrape_day(d):
                scraped += 1
            else:
                skipped += 1
        except Exception:
            log.exception("Failed to scrape %s", d)
            failed += 1
        d += timedelta(days=1)

    log.info(
        "Done. %d days total, %d scraped, %d skipped, %d failed",
        total_days, scraped, skipped, failed,
    )


def parse_date(s: str) -> date:
    return datetime.strptime(s, "%Y-%m-%d").date()


def main() -> None:
    parser = argparse.ArgumentParser(description="Scrape CZ electricity spot prices from OTE-CR")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--date", type=parse_date, help="Scrape a single date (YYYY-MM-DD)")
    group.add_argument("--yesterday", action="store_true", help="Scrape yesterday's data")
    group.add_argument("--from", dest="from_date", type=parse_date, help="Start of date range (YYYY-MM-DD)")
    parser.add_argument("--to", dest="to_date", type=parse_date, help="End of date range (YYYY-MM-DD), defaults to today")
    parser.add_argument("--verbose", "-v", action="store_true", help="Enable debug logging")

    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    if args.date:
        scrape_day(args.date)
    elif args.yesterday:
        scrape_day(date.today() - timedelta(days=1))
    elif args.from_date:
        end = args.to_date or date.today()
        scrape_range(args.from_date, end)


if __name__ == "__main__":
    main()
